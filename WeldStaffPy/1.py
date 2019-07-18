# -*- coding: utf-8 -*-
import os, sys
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter 

if __name__ == '__main__':
    workbook_ = load_workbook(u"2.xls")
    sheetnames =workbook_.get_sheet_names()
    print(sheetnames)
    